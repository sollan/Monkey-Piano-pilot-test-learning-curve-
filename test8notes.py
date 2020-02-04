'''
in case you're reading this, I didn't write the shitty part of this code
'''

##############Libraries##############
#library to generate the song
import pygame.midi
#Library to generate a GUI or use keyboard 
import pygame
#Other required libraries
from psychopy import parallel,core,event
import time
import random
import numpy
import datetime
import pickle
from openpyxl import Workbook
#Other python codes
# from Generate import Generate


#%%Function section
###############Function section for starting rhythmic notes###############
#One note played before each trial

# a dict that call each function 
# if you add more function adjust the dict to call them in the experiment

# def PD(lett):
#     if lett=='n1':
#         port.setData(2)
#     elif lett=='n2':
#         port.setData(3)
#     elif lett=='n3':
#         port.setData(4)
#     elif lett=='n4':
#         port.setData(5)

def RunExperiment(name = 's02_exp01', sequence = [60,62,60,65,62,64,65,62,64,65,62,60], 
                    MusicExp = 'no', filepath = ["./"]):



    
    #%%Variables
    ###############Variables##############
    #number of trials in one experiment 
    total_trials=15
    # initialize parallel port
    # port = parallel.ParallelPort(0xDC00)
    # Rythmic notes before each trial
    #1: one note
    num=1
    #Length of the sequence of notes
    N=8
    #Number of notes
    X=4
    #note sound
    #only if X=4, add more notes if X>4 and keyboard press ex: 'n5'
    # Change only this block and the loop keyboard block
    #pitch=[40,42,44,45]
    n1=60
    n2=62
    n3=64
    n4=65
    # name='s02_exp01'
    seq_notes=[n1,n2,n3,n4]
    string_notes=['n1','n2','n3','n4']

    #time of one note
    t1=.5
    portSetData = {'S1' : 'Start of Experiment', 'S2' : 'Listening N1',
                'S3' : 'Listening N2', 'S4' : 'Listening N3',
                'S5' : 'Listening N4', 'S6' : 'Reproducing N1','S7':'Reproducing N2','S8':'Reproducing N3','S9': 'Reproducing N4',
                'S10' : 'Listening beep', 'S11' : 'Reproducing Beep','S12':'correct note','S13':'wrong note',
                'S14' : 'End of Experiment'}

    #%% Experiment Name and Subject infog  
    DT=datetime.datetime.now().strftime("%y-%m-%d--%H:%M")
    # details of experiment output
    FilePath = filepath

    #%%Sequence generation
    #N.B : Make sure the length of the sequence and notes match the variables 
    # 'N' 
    #sequence of notes################## change a for custom sequence
        #############################################
    a=sequence # 60,62,64,65

    for i, note in enumerate(a):
        if note == 1:
            a[i] = n1
        if note == 2:
            a[i] = n2
        if note == 3:
            a[i] = n3
        if note == 4:
            a[i] = n4
    #############################################
    # a1=a
#required for printing the sequence in the excel sheet
    # for i in range(len(seq_notes)):
    #     a1=([ string_notes[i] if element==seq_notes[i] else element  for element in a1])
    # print(a1)
        
    #%%xlsx file and text file
    ###############xlxs file and text file###############
    # creating a workbook excel file
    wb=Workbook()
    #adding a sheet to the excel file
    worksheet1=wb.active
    #Naming the blocks and the excel file
    worksheet1['A1']='Experiment'
    worksheet1['B1']='Trials'
    worksheet1['C1']='Events'
    worksheet1['D1']='Note played'
    worksheet1['E1']='Note Replayed'
    worksheet1['F1']='keypress'
    worksheet1['G1']='Time relative'
    worksheet1['H1']='Time Absolute'
    worksheet1['I1']='B.Error'
    worksheet1['J1']='C.Error'
    personalInfo = {'name' : name[0:3], 'musicInstExp':MusicExp,'Seqnotes':seq_notes,'Seqvalue':string_notes, 'date':DT}
    with open(str(filepath)+str(name[0:3])+'.txt',"a") as text_file:
        text_file.write('\n')
        text_file.write("Personal information")
        text_file.write('\n')
    with open(filepath+personalInfo['name']+'.txt',"a") as text_file:
        for item in personalInfo:
            text_file.write("%s:%s\t" % (item, personalInfo[item]))
        text_file.write('\n')
    with open(filepath+name[0:3]+'.txt',"a")as text_file :
        text_file.write(str(a))
    events=total_trials*N
    for exper in range(events):
        worksheet1.cell(row=exper+2,column=1).value=name[7:9]
    PIK=name+'.dat'
    with open(PIK, "wb") as f:
        pickle.dump(str(a), f)

    #%%
    ###############Trial loop###############
    # port.setData(1)
    for tr in range(total_trials):
        time.sleep(3)
        #Initializing the MIDI object
        pygame.midi.init()
        #Setting the MIDI output
        player = pygame.midi.Output(2,buffer_size = 0)
        #Setting the MIDI instrument used
        player.set_instrument(0) # (0)piano (67--80) beeep sound somewhere


        def one():
            player.set_instrument(0) # (0)piano (67--80) beeep sound somewhere
            screen = pygame.display.set_mode([1, 1], 0) 
            print('replay the song')
            player.note_on(70, 127)
            time.sleep(t1)
            player.note_off(70, 127)
        #No note played before each trial
        def zero():
            print('replay the song')


        options={0:zero, 1:one}

        #Initializing the GUI screen required to run MIDI
        screen = pygame.display.set_mode([0, 0], 1)
        #pygame.key.set_repeat(1, 100)
        #Initializing the counters for time and error
        j=0
        o=0
        l=0
        i=0
        leng=N
        t=numpy.zeros(leng)
        toc=0
        tic=numpy.zeros(leng)
        dif=numpy.zeros(N)
        rel=numpy.zeros(N)
        x={}
        nots=numpy.zeros(N)
        counter=0
        performance=0
        count=N*tr+2
        #printing the trial number and events
        for cu in range(N):
            worksheet1.cell(row=count+cu,column=2).value=tr+1
            worksheet1.cell(row=count+cu,column=3).value=cu+1
        #Start of experiments
        print('Start of trial',tr+1) 
    #Playing the sound  
        #Rythm option
        # port.setData(10)
        options[num]()
        time.sleep(2)
        print('Now playing')
        #This loop runs the  sequence and print them inside th excel sheet
        for j in range(N):
            # PD(a1[j])
            player.note_on(a[j], 127)
            time.sleep(t1)
            player.note_off(a[j], 127)
            time.sleep(t1)
            worksheet1.cell(row=count+j,column=4).value=a[j]
    #------------------------------loop for keyboard----------------------------- 
    #Running the chosen rhythm
        time.sleep(1)
        # port.setData(11)
        options[num]()
        toc=time.time()
    #if there is more than 4 notes
    #the keyboard press need to entered manually as given in the commented part    
        while True:
            
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    raise SystemExit
                elif event.type == pygame.KEYDOWN:        
                    if event.key== pygame.K_g:
                        t[i]=time.time()
                        # port.setData(6)
                        player.note_on(n1, 127)
                        time.sleep(t1)
                        player.note_off(n1, 127)
                        x[i]=n1
                        # if x[i]==a[i]:
                        #     # port.setData(12)
                        # else:
                        #     # port.setData(13)
                        i=i+1
                        worksheet1.cell(row=((tr*N)+i)+1,column=6).value='k1'      
                    elif event.key == pygame.K_h:
                        t[i]=time.time()
                        # port.setData(7)
                        player.note_on(n2, 127)
                        time.sleep(t1)
                        player.note_off(n2, 127)
                        x[i]=n2
                        # if x[i]==a[i]:
                        #     # port.setData(12)
                        # else:
                        #     # port.setData(13)
                        i=i+1
                        worksheet1.cell(row=((tr*N)+i)+1,column=6).value='k2'            
                    elif event.key == pygame.K_j:
                        t[i]=time.time()
                        # port.setData(8)
                        player.note_on(n3, 127)
                        time.sleep(t1)
                        player.note_off(n3, 127)
                        x[i]=n3
                        # if x[i]==a[i]:
                        #     # port.setData(12)
                        # else:
                        #     # port.setData(13)
                        i=i+1
                        worksheet1.cell(row=((tr*N)+i)+1,column=6).value='k3'               
                    elif event.key ==pygame.K_k:
                        t[i]=time.time()
                        # port.setData(9)
                        player.note_on(n4, 127)
                        time.sleep(t1)
                        player.note_off(n4, 127)
                        x[i]=n4
                        # if x[i]==a[i]:
                        #     # port.setData(12)
                        # else:
                        #     # port.setData(13)
                        i=i+1
                        worksheet1.cell(row=((tr*N)+i)+1,column=6).value='k4'  
                    pygame.event.clear()
                elif event.type == pygame.KEYUP:
                    if event.key ==pygame.K_k:
                        time.sleep(.1)
                    elif event.key ==pygame.K_j:
                        time.sleep(.1)
                    elif event.key ==pygame.K_h:
                        time.sleep(.1)
                    elif event.key ==pygame.K_g:
                        time.sleep(.1)   
            
            if i==leng:
                break
        del player
        pygame.midi.quit()
        pygame.quit()
        
        print('trial end')
        
        
    #Calculate the time and the error and printing inside the excel sheet
        print(a)
        print(x)
        for l in range(N):
            if l==0 :
                dif[l]=t[0]-toc
            else:
                dif[l]=t[l]-t[(l-1)]
            if a[l] == x[l]:
                nots[l]=0
            else:
                nots[l]=1
                counter=counter+1
            worksheet1.cell(row=count+l,column=10).value=counter
            rel[l]=t[l]-toc
            worksheet1.cell(row=count+l,column=5).value=x[l]
            worksheet1.cell(row=count+l,column=7).value=dif[l]
            worksheet1.cell(row=count+l,column=8).value=rel[l]
            worksheet1.cell(row=count+l,column=9).value=nots[l]   
    wb.save(filepath+name+ ".xlsx")
    # port.setData(14)


if __name__ == "__main__":

    from test8notes import RunExperiment
    import numpy as np

    subject_name = 't08'
    MusicExp = 'no'
    filepath = './'

    sequence_list = ['01', '02', '03', '04', '05', 
                    '06', '07', '08', '09', '10', 
                    '11', '12', '13', '14', '15']
    sequence_name = [subject_name + '_exp' + item for item in sequence_list]

    with open('./input2.txt') as file:
        data = []
        for line in file:
            data.append(line)

    for i, sequence in enumerate(data):
        data[i] = sequence[:-1]

    sequence = []

    for item in data:
        for i in range(0, len(item), 3):
            sequence.append(int(item[i]))

    sequence = np.array(sequence).reshape(-1, 8)

    RunExperiment(sequence_name[0], sequence[-1], 'yes', filepath)
    RunExperiment(sequence_name[1], sequence[-2], 'yes', filepath)
    RunExperiment(sequence_name[2], sequence[-3], 'yes', filepath)